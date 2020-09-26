from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import Workbook
from openpyxl import load_workbook

import os
from os import path
import csv
import time
import urllib.request
import xlsxwriter


driver = webdriver.Chrome()
x = 2
url = "http://www.gmatfree.com/module-2/buses-passengers-and-folders/?explanation=1"
driver.get(url)
questions = driver.find_element_by_class_name('gfquestion')
question = questions.text

explanations = driver.find_element_by_class_name('gfexplanation')
explanation = explanations.text

alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
answer_pos = (explanation[-3])
answer_pos ="ans_" + repr(alphabet.index(answer_pos)+1)

answer_list = driver.find_element_by_class_name('gfchoices')
for element in answer_list.find_elements_by_tag_name('li'):
    radio = element.find_element_by_tag_name('input').get_attribute('value')
    if radio == answer_pos:
        label = element.find_element_by_tag_name('label')    
        img = label.find_element_by_tag_name('img').get_attribute("src")

filenames =url.split("/")
title = filenames[4]
filename ="module2/" + filenames[4] + ".png"
urllib.request.urlretrieve(img, filename)

book = Workbook()
sheet = book.active

sheet['A2'] = url
sheet['B2'] = title
sheet['C2'] = "The Free GMAT Prep Course"
sheet['D2'] = question
sheet['E2'] = filename
sheet['F2'] = explanation

book.save("demo2.xlsx")

nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
nextbtn = nextbtn0.find_element_by_tag_name('a')
time.sleep(0.5)
nextbtn.click()
time.sleep(1)


nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
nextbtn3 = nextbtn2.find_element_by_tag_name('a')
time.sleep(0.5)
nextbtn3.click()
url = driver.current_url
time.sleep(0.5)

while url != "http://www.gmatfree.com/module-3/ways-to-solve/":
    x = x + 1
    url = driver.current_url
    print(url)
    driver.get(url)
    questions = driver.find_element_by_class_name('gfquestion')
    question = questions.text

    explanations = driver.find_element_by_class_name('gfexplanation')
    explanation = explanations.text

    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    answer_pos = (explanation[-3])
    answer_pos ="ans_" + repr(alphabet.index(answer_pos)+1)
    
    answer_list = driver.find_element_by_class_name('gfchoices')
    for element in answer_list.find_elements_by_tag_name('li'):
        radio = element.find_element_by_tag_name('input').get_attribute('value')
        if radio == answer_pos:
            label = element.find_element_by_tag_name('label') 
            print(label.text)
            if label.text == '':   
                img = label.find_element_by_tag_name('img').get_attribute("src")
                filenames =url.split("/")
                title = filenames[4]
                filename ="module2/" + filenames[4] + ".png"
                urllib.request.urlretrieve(img, filename)
            else:
                filename = label.text

    rowA = "A" + repr(x)
    rowB = "B" + repr(x)
    rowC = "C" + repr(x)
    rowD = "D" + repr(x)
    rowE = "E" + repr(x)
    rowF = "F" + repr(x)

    book = load_workbook("demo2.xlsx")
    sheet = book.active

    sheet[rowA] = url
    sheet[rowB] = title
    sheet[rowC] = "The Free GMAT Prep Course"
    sheet[rowD] = question
    sheet[rowE] = filename
    sheet[rowF] = explanation

    book.save("demo2.xlsx")

    nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
    nextbtn = nextbtn0.find_element_by_tag_name('a')
    time.sleep(0.5)
    nextbtn.click()
    time.sleep(0.5)

    nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
    nextbtn3 = nextbtn2.find_element_by_tag_name('a')
    time.sleep(0.5)
    nextbtn3.click()
    time.sleep(0.5)
 
driver.quit()