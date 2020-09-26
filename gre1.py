from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.remote.webelement import WebElement
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from os import path
import csv
import time
import urllib.request
import xlsxwriter
#myquiz > table > tbody > tr:nth-child(1) > td > div.Bfquiz_plusIntro
driver = webdriver.Chrome()
x = 2
url = "https://www.test-guide.com/gre/free-gre-practice-tests/gre-math-multiple-choice-practice-test-1.html"
driver.get(url)
table = driver.find_element_by_css_selector("#myquiz > table > tbody > tr:nth-child(6) > th > div")

print(table.text)
# questions = driver.find_element_by_class_name('Bfquiz_plusQuestion')
# question = questions.text


# answer_lists = driver.find_element_by_class_name('bfradio')
# print(answer_lists)

# explanations = driver.find_element_by_class_name('Bfquiz_plusOptions')
# explanation = label.find_element_by_tag_name('img').get_attribute("src")

# searchpos = explanation.rfind("answer is")+11
# book = Workbook()
# sheet = book.active
# book.save("de.xlsx")

# if searchpos >= 12:

#     alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
#     answer_pos = (explanation[searchpos])
#     answer_pos ="ans_" + repr(alphabet.index(answer_pos)+1)

#     answer_list = driver.find_element_by_class_name('gfchoices')
#     for element in answer_list.find_elements_by_tag_name('li'):
#         radio = element.find_element_by_tag_name('input').get_attribute('value')
#         if radio == answer_pos:
#                 label = element.find_element_by_tag_name('label') 
#                 print(label.text)
#                 if label.text == '':   
#                     img = label.find_element_by_tag_name('img').get_attribute("src")
#                     filenames =url.split("/")
#                     title = filenames[4]
#                     filename ="module2/" + filenames[4] + ".png"
#                     urllib.request.urlretrieve(img, filename)
#                 else:
#                     filename = label.text
#                     filenames =url.split("/")
#                     title = filenames[4]

#     book = Workbook()
#     sheet = book.active

#     sheet['A2'] = url
#     sheet['B2'] = title
#     sheet['C2'] = "The Free GMAT Prep Course"
#     sheet['D2'] = question
#     sheet['E2'] = filename
#     sheet['F2'] = explanation

#     book.save("demo3.xlsx")

#     nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
#     nextbtn = nextbtn0.find_element_by_tag_name('a')
#     time.sleep(1)
#     nextbtn.click()
#     time.sleep(1)


#     nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
#     nextbtn3 = nextbtn2.find_element_by_tag_name('a')
#     time.sleep(1)
#     nextbtn3.click()
#     url = driver.current_url
#     time.sleep(1)
# else:
#     nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
#     nextbtn = nextbtn0.find_element_by_tag_name('a')
#     time.sleep(1)
#     nextbtn.click()
#     time.sleep(1)

#     nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
#     nextbtn3 = nextbtn2.find_element_by_tag_name('a')
#     time.sleep(1)
#     nextbtn3.click()
#     url = driver.current_url
#     time.sleep(1)

# while url != "http://www.gmatfree.com/module-112/nanoparticles-vi/?explanation=1":
#     x = x + 1
#     url = driver.current_url
#     print(url)
#     driver.get(url)
#     questions = driver.find_element_by_class_name('gfquestion')
#     question = questions.text

#     explanations = driver.find_element_by_class_name('gfexplanation')
#     explanation = explanations.text
#     searchpos = explanation.rfind("answer is")+11
#     if searchpos >= 12:

#         alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
#         answer_pos = (explanation[searchpos])
#         answer_pos ="ans_" + repr(alphabet.index(answer_pos)+1)    
    
#         answer_list = driver.find_element_by_class_name('gfchoices')
#         for element in answer_list.find_elements_by_tag_name('li'):
#             radio = element.find_element_by_tag_name('input').get_attribute('value')
#             if radio == answer_pos:
#                 label = element.find_element_by_tag_name('label') 
#                 print(label.text)
#                 if label.text == '':   
#                     img = label.find_element_by_tag_name('img').get_attribute("src")
#                     filenames =url.split("/")
#                     title = filenames[4]
#                     filename ="module2/" + filenames[4] + ".png"
#                     urllib.request.urlretrieve(img, filename)
#                 else:
#                     filenames =url.split("/")
#                     title = filenames[4]
#                     filename = label.text

#         rowA = "A" + repr(x)
#         rowB = "B" + repr(x)
#         rowC = "C" + repr(x)
#         rowD = "D" + repr(x)
#         rowE = "E" + repr(x)
#         rowF = "F" + repr(x)

#         book = load_workbook("demo3.xlsx")
#         sheet = book.active

#         sheet[rowA] = url
#         sheet[rowB] = title
#         sheet[rowC] = "The Free GMAT Prep Course"
#         sheet[rowD] = question
#         sheet[rowE] = filename
#         sheet[rowF] = explanation

#         book.save("demo3.xlsx")

#         nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
#         nextbtn = nextbtn0.find_element_by_tag_name('a')
#         time.sleep(1)
#         nextbtn.click()
#         time.sleep(1)

#         nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
#         nextbtn3 = nextbtn2.find_element_by_tag_name('a')
#         time.sleep(1)
#         nextbtn3.click()
#         time.sleep(1)
    
#     else:
#         nextbtn0 = driver.find_element_by_class_name('wpcw_fe_navigation_box')
#         nextbtn = nextbtn0.find_element_by_tag_name('a')
#         time.sleep(1)
#         nextbtn.click()
#         time.sleep(1)

#         nextbtn2 = driver.find_element_by_class_name('wpcw_fe_quiz_submit')
#         nextbtn3 = nextbtn2.find_element_by_tag_name('a')
#         time.sleep(1)
#         nextbtn3.click()
#         time.sleep(1)

driver.quit()